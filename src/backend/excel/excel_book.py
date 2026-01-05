from pathlib import Path

from openpyxl import load_workbook

from backend.excel.excel_sheet import ExcelSheet
from backend.excel.excel_sheet_equality import excelsheet_equals
from logger import logger
from logs_label import ExcelNotExisting, NoRightWorksheet


class ExcelBook:
    """Wrapper around openpyxl WorkBook"""

    def __init__(self, path_excel: Path):
        self.path_excel = path_excel
        if not path_excel.exists():
            raise ExcelNotExisting(path_excel)

        self.wb = load_workbook(path_excel, rich_text=True)
        self.first_es: ExcelSheet = ExcelSheet(
            ws=self.wb.worksheets[0], name=self.wb.sheetnames[0], wb=self.wb
        )
        self.save = self.wb.save

    def get_excel_sheet(self, name: str) -> ExcelSheet:
        if name not in self.wb.sheetnames:
            raise NoRightWorksheet(
                excel_name=self.get_excel_name(),
                page_name=f"The page worksheet named '{name}' does not exist or is does not have the right name.",
            )
        return ExcelSheet(
            self.wb.worksheets[self.wb.sheetnames.index(name)], name=name, wb=self.wb
        )

    def get_excel_name(self) -> str:
        return Path(self.path_excel).name

    def equals(self, other: "ExcelBook") -> bool:
        if set(self.wb.sheetnames) != set(other.wb.sheetnames):
            logger.info("Excel equality : not same sheetnames")
            return False

        return all(
            excelsheet_equals(self.get_excel_sheet(name), other.get_excel_sheet(name))
            for name in self.wb.sheetnames
        )

    def __iter__(self):
        return iter(
            ExcelSheet(ws, name=ws_name, wb=self.wb)
            for ws, ws_name in zip(self.wb.worksheets, self.wb.sheetnames)
        )
