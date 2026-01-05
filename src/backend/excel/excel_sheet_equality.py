from copy import copy
from itertools import product
from xml.etree.ElementTree import fromstring

from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.styles import Border
from openpyxl.styles.borders import Border
from openpyxl.styles.colors import Color
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter

from backend.excel.cell import Cell
from backend.excel.excel_sheet import ExcelSheet
from logger import f, logger

# Excel theme color order (fixed by spec)
THEME_COLOR_ORDER = [
    "lt1",
    "dk1",
    "lt2",
    "dk2",
    "accent1",
    "accent2",
    "accent3",
    "accent4",
    "accent5",
    "accent6",
    "hlink",
    "folHlink",
]


def get_theme_colors(wb):
    """
    Returns a list of RGB strings (RRGGBB) in theme index order
    """
    if not wb.loaded_theme:
        return None

    return [
        "FFFFFF",
        "000000",
        "FFFFFF",
        "000000",
        "4F81BD",
        "C0504D",
        "9BBB59",
        "8064A2",
        "4BACC6",
        "F79646",
        "0000FF",
        "800080",
    ]
    root = fromstring(wb.loaded_theme)

    # Find color scheme
    ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
    scheme = root.find(".//a:clrScheme", ns)
    if scheme is None:
        return None

    colors = []
    for name in THEME_COLOR_ORDER:
        el = scheme.find(f"a:{name}", ns)
        if el is None:
            colors.append(None)
            continue

        srgb = el.find("a:srgbClr", ns)
        sysc = el.find("a:sysClr", ns)

        if srgb is not None:
            colors.append(srgb.attrib["val"])
        elif sysc is not None:
            colors.append(sysc.attrib["lastClr"])
        else:
            colors.append(None)

    logger.info(colors)

    return colors


def apply_tint(rgb, tint):
    # rgb: 'RRGGBB'
    r = int(rgb[0:2], 16)
    g = int(rgb[2:4], 16)
    b = int(rgb[4:6], 16)

    def tint_channel(c):
        if tint < 0:
            return int(c * (1 + tint))
        return int(c + (255 - c) * tint)

    r = tint_channel(r)
    g = tint_channel(g)
    b = tint_channel(b)

    return f"FF{r:02X}{g:02X}{b:02X}"


def normalize_color_visual(color, wb):
    if color is None:
        return None

    # RGB → return
    if color.__dict__.get("rgb"):
        return color.rgb

    # Theme → resolve
    if color.theme is not None:
        theme_colors = get_theme_colors(wb)
        if not theme_colors:
            return None

        base = theme_colors[color.theme]
        if base is None:
            return None

        return apply_tint(base, color.tint or 0.0)

    # Indexed (optional)
    if color.indexed is not None:
        return color.indexed

    if color.auto:
        return "auto"

    return None


def side_equal(s1, s2, wb):
    if s1 is s2:
        return True
    if s1 is None or s2 is None:
        return False

    return s1.style == s2.style and normalize_color_visual(
        s1.color, wb
    ) == normalize_color_visual(s2.color, wb)


def equals_border(b1: Border, b2: Border, wb) -> bool:
    return (
        b1.outline == b2.outline
        and b1.diagonalUp == b2.diagonalUp
        and b1.diagonalDown == b2.diagonalDown
        and b1.start == b2.start
        and b1.end == b2.end
        and side_equal(b1.left, b2.left, wb)
        and side_equal(b1.right, b2.right, wb)
        and side_equal(b1.top, b2.top, wb)
        and side_equal(b1.bottom, b2.bottom, wb)
        and side_equal(b1.diagonal, b2.diagonal, wb)
        and side_equal(b1.vertical, b2.vertical, wb)
        and side_equal(b1.horizontal, b2.horizontal, wb)
    )


def equals_fill(f1: PatternFill, f2: PatternFill) -> bool:
    return f1.fgColor == f2.fgColor and f1.bgColor == f2.bgColor


def equals_fonts(f1: Font, f2: Font, wb) -> bool:
    if f1 is f2:
        return True
    if f1 is None or f2 is None:
        return False

    return (
        f1.name == f2.name
        and f1.sz == f2.sz
        and f1.b == f2.b
        and f1.i == f2.i
        and f1.u == f2.u
        and f1.strike == f2.strike
        and normalize_color_visual(f1.color, wb) == normalize_color_visual(f2.color, wb)
        and f1.vertAlign == f2.vertAlign
        and f1.scheme == f2.scheme
    )


def equals_text_block(tb1: TextBlock, tb2: TextBlock, wb) -> bool:

    if tb1.text != tb2.text:
        logger.info(
            f"Excel equality : CellRichText not same text : {tb1.text} != {tb2.text}"
        )
        return False

    if not equals_fonts(tb1.font, tb2.font, wb):
        logger.info(
            f"Excel equality : CellRichText not same font : {tb1.font}\n!=\n{tb2.font}"
        )
        return False

    return True


def equals_rich_text(rt1: CellRichText, rt2: CellRichText, wb):
    if len(rt1) != len(rt2):
        logger.info(
            f"Excel equality : CellRichText not same len : {len(rt1)} != {len(rt2)}"
        )
        return False

    for run1, run2 in zip(rt1, rt2):

        if type(run1) != type(run2):
            logger.info(
                f"Excel equality : CellRichText not same element type : {type(run1)} != {type(run2)}"
            )
            return False

        if isinstance(run1, TextBlock):
            if not equals_text_block(run1, run2, wb):
                return False

        elif isinstance(run1, str):
            if run1 != run2:
                return False

        else:
            raise RuntimeError(f"Type not handled : {type(run1)}")

    return True


def equals_cell(c1: Cell, c2: Cell, wb) -> bool:

    # second_cell
    # if c1.row != 1 or c1.col != 1:
    #     return True

    # text
    if c1.str != c2.str:
        logger.info(f"Excel equality : str not equal : '{c1.str}' != '{c2.str}'")
        return False

    # font
    if c1.str and not equals_fonts(c1.font, c2.font, wb):
        logger.info(f"Excel equality : font not equal : {c1.font}\n!=\n{c2.font}")
        logger.info(
            f"Excel equality : colors : c1='{normalize_color_visual(c1.font.color, wb)}' ; c2='{normalize_color_visual(c2.font.color, wb)}'"
        )
        return False

    # fill
    if c1.str and not equals_fill(c1.fill, c2.fill):
        logger.info(f"Excel equality : fill not equal : {c1.fill}\n!=\n{c2.fill}")
        return False

    # border
    if not equals_border(c1.border, c2.border, wb):
        logger.info(f"Excel equality : border not equal : {c1.border}\n!=\n{c2.border}")
        return False

    # value
    v1 = c1.value
    v2 = c2.value

    if type(v1) != type(v2):
        logger.info(f"Excel equality : Not same type : {type(v1)} != {type(v2)}")
        return False

    # nothing, str, int
    if v1 is None or isinstance(v1, str) or isinstance(v1, int):
        return True

    # cell rich text
    assert isinstance(v1, CellRichText) and isinstance(v2, CellRichText)
    return equals_rich_text(v1, v2, wb)


def is_empty(
    es: ExcelSheet, row_min: int, row_max: int, col_min: int, col_max: int
) -> bool:
    for row, col in product(
        range(row_min, row_max + 1),
        range(col_min, col_max + 1),
    ):
        if not es.get_text_cell(row=row, col=col) is None:
            logger.info(
                f"Excel equality : Sheet '{es.name}' : cell {f(row=row, col=col)} should be empty, here is the content : '{es.get_text_cell(row=row, col=col)}'."
            )
            return False
    return True


def excelsheet_equals(es1: ExcelSheet, es2: ExcelSheet) -> bool:

    min_dim_row = min(es1.get_row_dimension(), es2.get_row_dimension())
    min_dim_col = min(es1.get_col_dimension(), es2.get_col_dimension())

    # check ranges
    if es1.ws.merged_cells.ranges != es2.ws.merged_cells.ranges:
        logger.info(
            f"Excel equality : Sheets {f(name1=es1.name, name2=es2.name)} : ranges not equal {f(left=es1.ws.merged_cells.ranges, right=es2.ws.merged_cells.ranges)}"
        )
        return False

    # check cells equality
    for row, col in product(range(1, min_dim_row + 1), range(1, min_dim_col + 1)):

        c1 = es1.get_cell(row, col)
        c2 = es2.get_cell(row, col)

        if not equals_cell(c1, c2, es1.wb):
            logger.info(
                f"Excel equality : Sheets {f(name1=es1.name, name2=es2.name)} : cells not equal {f(row=row, col=col)}"
            )
            return False

    # check emptyness
    return (
        is_empty(
            es=es1,
            row_min=min_dim_row + 1,
            row_max=es1.get_row_dimension(),
            col_min=1,
            col_max=es1.get_col_dimension(),
        )
        and is_empty(
            es=es1,
            row_min=1,
            row_max=es1.get_row_dimension(),
            col_min=min_dim_col + 1,
            col_max=es1.get_col_dimension(),
        )
        and is_empty(
            es=es2,
            row_min=min_dim_row + 1,
            row_max=es2.get_row_dimension(),
            col_min=1,
            col_max=es2.get_col_dimension(),
        )
        and is_empty(
            es=es2,
            row_min=1,
            row_max=es2.get_row_dimension(),
            col_min=min_dim_col + 1,
            col_max=es2.get_col_dimension(),
        )
    )
