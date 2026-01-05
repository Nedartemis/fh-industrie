from copy import copy
from typing import Any, Callable, List, Optional, Tuple

from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.styles.fonts import Font
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from backend.excel.cell import Cell
from backend.table.table_base import TableBase
from logger import WARNING, f, logger
from logs_label import EmptynessExcelCell, ExactnessExcelCell, FullnessExcelCell


class ExcelSheet(Cell, TableBase):
    """Wrapper around openpyxl Worksheet"""

    def __init__(self, ws: Worksheet, name: str, wb: Workbook):
        self.ws = ws
        self.name = name
        self.wb = wb

    # ------------------- Getter -------------------

    @staticmethod
    def none_transformation(s: Optional[Any]) -> Optional[Any]:
        return None if (s is None or s == "" or s == "None") else s

    def get_cell(self, row: int, col: int) -> Cell:

        c = self.ws.cell(row, col)
        # assert isinstance(font, Font), f"type : {type(font)}"

        return Cell(
            row=row,
            col=col,
            value=ExcelSheet.none_transformation(self.ws.cell(row, col).value),
            str=self.get_text_cell(row, col),
            font=c.font,
            fill=c.fill,
            border=c.border,
            alignment=c.alignment,
            number_format=c.number_format,
            protection=c.protection,
            has_style=c.has_style,
        )

    def get_text_cell(self, row: int, col: int) -> Optional[str]:
        s = str(self.ws.cell(row, col).value).strip()
        return ExcelSheet.none_transformation(s)

    def get_row_dimension(self) -> int:
        return self.ws.max_row

    def get_col_dimension(self) -> int:
        return self.ws.max_column

    # ------------------- Modifiers -------------------

    def insert_rows(self, row: int, amount: int) -> None:
        if amount == 0:
            return

        self.ws.insert_rows(row, amount=amount)

    def replace_text_in_cell(
        self,
        row: int,
        col: int,
        replace_text: Callable[[str], Tuple[str, int]],
    ) -> int:
        nb_changes = 0

        # read
        cell_value = self.ws.cell(row, col).value
        if cell_value is None:
            return nb_changes

        if isinstance(cell_value, str):
            # replace
            cell_value, nb_new_changes = replace_text(cell_value)
            nb_changes += nb_new_changes

            # write
            self.ws.cell(row, col, value=cell_value)
        elif isinstance(cell_value, CellRichText):
            # replace
            for i, e in enumerate(cell_value):
                text = e if isinstance(e, str) else e.text

                text, nb_new_changes = replace_text(text)
                nb_changes += nb_new_changes

                if isinstance(e, str):
                    cell_value[i] = text
                else:
                    cell_value[i] = TextBlock(e.font, text)

                self.ws.cell(row, col, value=cell_value)
        else:
            raise ValueError(f"Excel cell type not implemmented : {type(cell_value)}")

        return nb_changes

    def erase_cell(self, row: int, col: int) -> None:
        cell = self.ws.cell(row=row, column=col)
        cell.value = ""
        cell.font = Font()

    # ------------------- Checkers -------------------

    def check_content_cell(
        self,
        page_name: str,
        row: int,
        col: int,
        expected_content: str,
        log_level: str = WARNING,
    ) -> bool:
        """
        Returns:
            bool: True if an error has been detected
        """

        content = self.get_text_cell(row=row, col=col)
        if content is None or content.lower() != expected_content.lower():
            logger.log(
                level=log_level,
                msg=f"{page_name} : Text cell{f(row=row, col=col)} should be '{expected_content}' but is '{content}'",
                extra=ExactnessExcelCell(
                    page_name=page_name,
                    row=row,
                    col=col,
                    expected=expected_content,
                    actual=content,
                ),
            )
            return True
        return False

    def check_emptiness_row(
        self,
        page_name: str,
        row: int,
        header_name_cols: Tuple[str, int],
        log_level: str = WARNING,
    ) -> bool:
        """
        Returns:
            bool: True if an error has been detected
        """

        not_empty_cols = [
            (header_name, content)
            for header_name, col in header_name_cols
            if (content := self.get_text_cell(row=row, col=col))
        ]

        if not_empty_cols:
            not_empty_cols_names = [header_name for header_name, _ in not_empty_cols]
            contents = [content for _, content in not_empty_cols]
            logger.log(
                level=log_level,
                msg=f"{page_name} : Column(s) [{', '.join(not_empty_cols_names)}] is/are not empty on the row {row}.\n"
                + f"Here is the content : {contents}.\n"
                + "The excel may have a mistake.",
                extra=FullnessExcelCell(
                    page_name=page_name,
                    header_names=not_empty_cols,
                    row=row,
                    actuals=contents,
                ),
            )

        return not_empty_cols

    def check_fullness_row(
        self,
        page_name: str,
        row: int,
        header_name_cols: List[Tuple[str, int]],
        log_level: str = WARNING,
    ) -> bool:
        """
        Returns:
            bool: True if an error has been detected
        """

        empty_cols = [
            header_name
            for header_name, col in header_name_cols
            if self.get_text_cell(row=row, col=col) is None
        ]

        if empty_cols:
            logger.log(
                level=log_level,
                msg=f"{page_name} : Column(s) [{', '.join(empty_cols)}] is/are empty on the row {row}.\n"
                + f"The excel may have a mistake.",
                extra=EmptynessExcelCell(
                    page_name=page_name, header_names=empty_cols, row=row
                ),
            )
        return empty_cols

    # ------------------- Copy -------------------

    def copy_cell(self, src_cell: Cell, row: int, col: int) -> None:

        dst_cell = self.ws.cell(row=row, column=col)

        # Value (str, number, or CellRichText)
        if isinstance(src_cell.value, CellRichText):
            dst_cell.value = copy(src_cell.value)
        else:
            dst_cell.value = src_cell.value

        # Style objects (MUST use copy)
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.border = copy(src_cell.border)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy(src_cell.protection)


if __name__ == "__main__":
    from backend.excel.excel_book import ExcelBook
    from vars import PATH_TEST_DOCS, PATH_TEST_DOCS_TESTSUITE

    path = PATH_TEST_DOCS_TESTSUITE / "excel" / "equals" / "merged_cell.xlsx"

    from openpyxl.cell.cell import MergedCell

    es = ExcelBook(path).first_es
    print(type(es.ws.cell(1, 1)))
    c = es.ws.cell(1, 2)
    assert isinstance(c, MergedCell)
    for e in es.ws.merged_cells.ranges:
        print(e, type(e))
    print(type(es.ws.cell(2, 1)))
