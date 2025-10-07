import itertools
from typing import Dict, List, Optional, Tuple

import openpyxl.cell.rich_text
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.worksheet.worksheet import Worksheet

from backend.generation import BORDER_LEFT, BORDER_RIGHT, HARMONIZE_LABEL_INFO
from backend.generation.replace_text import replace_text


class ExcelManager:

    def __init__(self, path_excel):
        self.wb = load_workbook(path_excel, rich_text=True)
        self.ws: Worksheet = self.wb.worksheets[0]

    # ------------------- Public Method -------------------

    @classmethod
    def get_text(cls, ws: Worksheet, row: int, col: int) -> Optional[str]:
        s = cls.get_text_cell(ws.cell(row, col).value).strip()
        return None if not s or s == "None" else s

    @staticmethod
    def get_text_cell(cell_value) -> str:
        return str(cell_value)

    def get_worksheet(self, name: str) -> Worksheet:
        if name not in self.wb.sheetnames:
            raise RuntimeError(
                f"The page worksheet named '{name}' does not exist or is does not have the right name."
            )
        return self.wb.worksheets[self.wb.sheetnames.index(name)]

    def replace_content(self, infos: dict) -> None:

        ws = self.ws
        n = 100

        nb_changes = 0

        # instruction column
        print(self.get_text(ws, 1, 1))
        has_column_instruction = self.get_text(ws, 1, 1) == "colonne_instruction"
        column_instruction = None
        first_column = 2 if has_column_instruction else 2
        if has_column_instruction:
            # read column instruction
            column_instruction = [self.get_text(ws, row, col=1) for row in range(1, n)]
            # delete column instruction
            for row in range(1, n):
                ws.cell(row, 1, value="")

        # independent info

        # -- filter
        independent_infos = {
            info_name: value
            for info_name, value in infos.items()
            if isinstance(value, str)
        }

        print(independent_infos)

        # -- replace
        for row, col in itertools.product(range(1, n), range(1, n)):
            nb_changes += self._replace_cell(ws, row, col, independent_infos)

        print(f"Excel changes indenpendent : {nb_changes}")

        # list info

        if not column_instruction:
            return nb_changes

        # -- filter
        list_infos = {
            info_name: value
            for info_name, value in infos.items()
            if isinstance(value, list)
        }

        # -- read instructions
        lists_rows: Dict[str, Tuple[str, str]] = {}
        name_list = None
        for idx_row, text in enumerate(column_instruction):
            if text is None or not ":" in text:
                continue

            instruction, name_list_current = text.split(":")
            print(">", instruction, name_list_current)
            if name_list != name_list_current:
                assert instruction == "debut_liste"
                name_list = name_list_current
                lists_rows[name_list] = (idx_row + 1, None)
            else:
                assert instruction == "fin_liste"
                lists_rows[name_list] = (lists_rows[name_list][0], idx_row + 1)

        print(lists_rows)

        # -- replace
        for list_name in lists_rows.keys():

            start_row, end_row = lists_rows[list_name]

            infos = list_infos.get(list_name)
            print("list_name :", list_name, ";", "infos :", infos)
            if infos is None:
                continue

            # insert rows
            nb_rows_list = end_row - start_row + 1
            nb_to_add = nb_rows_list * (len(infos) - 1)
            ws.insert_rows(end_row + 1, amount=nb_to_add)
            print("nb_to_add :", nb_to_add)

            # copy content
            texts: List[List[str]] = [
                [ws.cell(row, col).value for col in range(first_column, 10)]
                for row in range(start_row, end_row + 1)
            ]

            for idx in range(1, len(infos)):
                for offset_row, lst in enumerate(texts):
                    row = start_row + nb_rows_list * idx + offset_row

                    for col, cell_value in enumerate(lst, start=first_column):
                        cell_value_copy = (
                            cell_value
                            if not isinstance(cell_value, CellRichText)
                            else CellRichText(cell_value.copy())
                        )
                        if cell_value_copy:
                            print(cell_value, cell_value_copy)
                            ws.cell(row, col, value=cell_value_copy)

            # update other start and end
            for list_name_c in lists_rows.keys():
                start_row_c, end_row_c = lists_rows[list_name_c]
                lists_rows[list_name_c] = start_row_c + nb_to_add, end_row_c + nb_to_add

            # fill row
            for idx_element, infos_one_element in enumerate(infos):
                infos_one_element = {
                    f"{list_name}:{sub_name}": value
                    for sub_name, value in infos_one_element.items()
                }
                print("infos_one_element :", infos_one_element)

                offset = idx_element * nb_rows_list
                for row, col in itertools.product(
                    range(start_row + offset, end_row + offset + 1),
                    range(first_column, n),
                ):
                    nb_changes += self._replace_cell(ws, row, col, infos_one_element)

        print(f"Excel changes list : {nb_changes}")

    def save(self, path_excel) -> None:
        self.wb.save(path_excel)

    # ------------------- Private Method -------------------

    @staticmethod
    def _change(input: str, infos: Dict[str, str], count: int) -> Tuple[str, int]:
        output, plus_count = replace_text(
            s=input,
            pair_old_new=list(infos.items()),
            border_left=BORDER_LEFT,
            border_right=BORDER_RIGHT,
            do_harmonization=HARMONIZE_LABEL_INFO,
        )
        return output, count + plus_count

    @staticmethod
    def _replace_cell(ws: Worksheet, row: int, col: int, infos: Dict[str, str]) -> int:
        nb_changes = 0

        # read
        cell_value = ws.cell(row, col).value
        if cell_value is None:
            return nb_changes

        if isinstance(cell_value, str):
            # replace
            cell_value, nb_changes = ExcelManager._change(cell_value, infos, nb_changes)

            # write
            ws.cell(row, col, value=cell_value)
        elif isinstance(cell_value, openpyxl.cell.rich_text.CellRichText):
            # replace
            for i, e in enumerate(cell_value):
                text = e if isinstance(e, str) else e.text

                text, nb_changes = ExcelManager._change(text, infos, nb_changes)

                if isinstance(e, str):
                    cell_value[i] = text
                else:
                    cell_value[i] = TextBlock(e.font, text)

                ws.cell(row, col, value=cell_value)
        else:
            raise ValueError(f"Excel cell type not implemmented : {type(cell_value)}")

        return nb_changes


def _main():
    from vars import PATH_TEST

    # check if the conversion of CellRichText to str is good
    em = ExcelManager(PATH_TEST / "test_celine" / "fichier_configuration.xlsx")
    ws = em.get_worksheet("Infos à extraire")
    for row in range(1, len(ws.row_dimensions)):
        cell_value = ws.cell(row, column=2).value
        s = em.get_text_cell(cell_value)
        print(type(cell_value), cell_value, s)


if __name__ == "__main__":
    _main()
